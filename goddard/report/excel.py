"""Generated Excel report -- READ ONLY.

Spec section 9.

The workbook is an output, not an input. It is regenerated on every run and
editing it changes nothing. All inputs live in ``goddard/config/``.

This is the deliberate inversion of the 25-26 model, where the spreadsheet was
simultaneously the input surface, the solver and the report -- which is how a
hard-coded launch altitude of 524 m could sit unnoticed beside a stated 1255 m.
"""

from __future__ import annotations

from pathlib import Path

from goddard.band import BandResult
from goddard.sim import FlightResult

_M_TO_FT = 1.0 / 0.3048


def _style_header(ws, row: int = 1) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="1F3864")
    for cell in ws[row]:
        if cell.value is not None:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws, max_width: int = 40) -> None:
    from openpyxl.utils import get_column_letter

    for i, column in enumerate(ws.iter_cols(), start=1):
        width = max(
            (len(str(c.value)) for c in column[:200] if c.value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, max_width)


def write(
    result: FlightResult,
    path: str | Path,
    band: BandResult | None = None,
    sample_stride: int = 10,
) -> Path:
    """Write the flight report workbook.

    Parameters
    ----------
    sample_stride : write every Nth trajectory row. At dt = 0.01 s a full flight
        is ~25,000 rows, which Excel handles poorly and nobody reads; 10 keeps
        0.1 s resolution.
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for the Excel report. "
            "Install with: pip install -e '.[report]'"
        ) from exc

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ------------------------------------------------------------- Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["GODDARD 26-27 FLIGHT REPORT"])
    ws.append(["GENERATED FILE -- DO NOT EDIT. Inputs live in goddard/config/."])
    ws.append([])
    ws.append(["Metric", "Value", "Units"])

    rows = [
        ("Apogee", result.apogee_agl_m, "m AGL"),
        ("Apogee", result.apogee_ft, "ft AGL"),
        ("Target", 50000.0, "ft AGL"),
        ("Target met", "YES" if result.apogee_ft >= 50000.0 else "NO", ""),
        ("Max Mach", result.max_mach, ""),
        ("Max speed", result.max_speed_ms, "m/s"),
        ("Max dynamic pressure", result.max_dynamic_pressure_Pa / 1000.0, "kPa"),
        ("Max acceleration", result.max_acceleration_g, "g"),
        ("Rail exit velocity", result.rail_exit_velocity_ms, "m/s"),
        ("Min static margin", result.min_static_margin, "calibers"),
        ("Min web remaining", result.min_web_fraction * 100.0, "%"),
        ("Min chug margin", result.min_chug_margin, "x criterion"),
        ("Peak tip temperature", result.peak_tip_temperature_K, "K"),
    ]
    for name, value, unit in rows:
        ws.append([name, value, unit])

    ws.append([])
    ws.append(["Warnings"])
    for w in result.warnings:
        ws.append([w])
    if result.terminated_reason:
        ws.append([f"Terminated: {result.terminated_reason}"])
    _style_header(ws, row=4)
    _autosize(ws)

    # ------------------------------------------------------------- Events
    ws = wb.create_sheet("Events")
    ws.append(["Event", "Time (s)", "Altitude AGL (m)", "Altitude (ft)",
               "Velocity (m/s)", "Mach"])
    for record in result.events.all():
        ws.append([
            record.event.value,
            record.time_s,
            record.altitude_m,
            record.altitude_m * _M_TO_FT,
            record.velocity_ms,
            record.mach,
        ])
    _style_header(ws)
    _autosize(ws)

    # --------------------------------------------------------- Trajectory
    ws = wb.create_sheet("Trajectory")
    ws.append([
        "Time (s)", "Altitude AGL (m)", "Altitude (ft)", "Downrange (m)",
        "Speed (m/s)", "Mach", "Accel (g)", "Mass (kg)",
        "Dynamic pressure (kPa)", "C_D", "Static margin (cal)",
        "Roll rate (rad/s)", "Alpha (deg)",
    ])
    import math
    for s in result.samples[::sample_stride]:
        ws.append([
            s.t, s.altitude_agl, s.altitude_agl * _M_TO_FT, s.x,
            s.speed, s.mach, s.acceleration_g, s.mass,
            s.dynamic_pressure / 1000.0, s.cd, s.static_margin,
            s.roll_rate, math.degrees(s.alpha),
        ])
    _style_header(ws)
    _autosize(ws)

    # -------------------------------------------------------------- Motor
    ws = wb.create_sheet("Motor")
    ws.append([
        "Time (s)", "Thrust (N)", "Chamber pressure (MPa)", "O/F ratio",
        "Web remaining (%)", "Chug margin",
    ])
    for s in result.samples[::sample_stride]:
        if s.thrust <= 0.0:
            continue
        ws.append([
            s.t, s.thrust, s.chamber_pressure / 1e6, s.of_ratio,
            s.web_fraction * 100.0, s.chug_margin,
        ])
    _style_header(ws)
    _autosize(ws)

    # ------------------------------------------------------- Band envelope
    if band is not None and band.envelopes:
        ws = wb.create_sheet("Band Envelope")
        ws.append(["Metric", "Worst", "Best", "Driving corner", "Note"])
        for e in band.envelopes:
            ws.append([
                e.metric, e.worst, e.best, e.driving_corner,
                e.conservative_direction,
            ])
        ws.append([])
        ws.append([
            "Conservative has no single direction -- see spec section 6.1. "
            "Lower regression means HIGHER O/F, and the burnthrough case lives "
            "at HIGH regression."
        ])
        ws.append([])
        ws.append(["Corner", "Apogee (ft)", "Max Mach", "Min web (%)", "Status"])
        for c in band.corners:
            if c.ok:
                ws.append([
                    c.label(), c.result.apogee_ft, c.result.max_mach,
                    c.result.min_web_fraction * 100.0, "ok",
                ])
            else:
                ws.append([c.label(), None, None, None, c.error])
        _style_header(ws)
        _autosize(ws)

    wb.save(path)
    return path
